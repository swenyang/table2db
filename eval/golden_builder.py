"""Generate golden standard SQLite DBs by analyzing Excel files.

Each file gets a manually-crafted ideal DB reflecting the correct
interpretation of the Excel structure.
"""
import os
import sqlite3
import json
import openpyxl
from typing import Any


EVAL_DIR = os.path.dirname(__file__)
SOURCE_DIRS = {
    "references": os.path.join(EVAL_DIR, "source_files", "references"),
    "deliverables": os.path.join(EVAL_DIR, "source_files", "deliverables"),
}
GOLDEN_DIRS = {
    "references": os.path.join(EVAL_DIR, "golden_dbs", "references"),
    "deliverables": os.path.join(EVAL_DIR, "golden_dbs", "deliverables"),
}


def create_golden_db(db_path: str, tables: dict[str, dict]):
    """Create a golden standard SQLite DB.
    
    Args:
        db_path: Output .db file path.
        tables: {
            "table_name": {
                "columns": [{"name": str, "type": str}, ...],
                "rows": [[val, ...], ...],
                "source_sheet": str,
                "description": str,
            }
        }
    """
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    if os.path.exists(db_path):
        os.unlink(db_path)
    
    conn = sqlite3.connect(db_path)
    for tbl_name, tbl_def in tables.items():
        cols = tbl_def["columns"]
        col_sql = ", ".join(f'"{c["name"]}" {c["type"]}' for c in cols)
        conn.execute(f'CREATE TABLE "{tbl_name}" ({col_sql})')
        
        if tbl_def.get("rows"):
            placeholders = ", ".join(["?"] * len(cols))
            conn.executemany(
                f'INSERT INTO "{tbl_name}" VALUES ({placeholders})',
                [tuple(r) for r in tbl_def["rows"]],
            )
    
    # Store metadata
    conn.execute("CREATE TABLE _golden_meta (key TEXT, value TEXT)")
    for tbl_name, tbl_def in tables.items():
        conn.execute("INSERT INTO _golden_meta VALUES (?, ?)",
                     (f"table:{tbl_name}:source_sheet", tbl_def.get("source_sheet", "")))
        conn.execute("INSERT INTO _golden_meta VALUES (?, ?)",
                     (f"table:{tbl_name}:description", tbl_def.get("description", "")))
        conn.execute("INSERT INTO _golden_meta VALUES (?, ?)",
                     (f"table:{tbl_name}:row_count", str(len(tbl_def.get("rows", [])))))
    
    conn.commit()
    conn.close()


def read_excel_raw(path: str) -> dict:
    """Read Excel file and return raw structure for analysis."""
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {}
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        result[ws.title] = {
            "rows": rows,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "merged": [str(m) for m in ws.merged_cells.ranges],
        }
    return result
