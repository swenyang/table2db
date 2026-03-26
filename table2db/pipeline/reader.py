"""Stage 1: Read Excel/CSV/TSV files into WorkbookData."""
from __future__ import annotations

import csv
import io
import os
from typing import Any, BinaryIO, Union

import openpyxl
from openpyxl.utils import get_column_letter

from table2db.errors import FileReadError, UnsupportedFormatError
from table2db.models import SheetData, WorkbookData

_ERROR_VALUES = {"#REF!", "#N/A", "#DIV/0!", "#VALUE!", "#NAME?", "#NULL!", "#NUM!"}


def read_workbook(
    source: Union[str, BinaryIO],
    skip_hidden_sheets: bool = True,
    file_name: str | None = None,
    color_mode: str = "ignore",
) -> WorkbookData:
    """Read an Excel/CSV/TSV file and return raw WorkbookData.

    Args:
        source: File path (str) or file-like object (BytesIO, UploadFile.file, etc.)
        skip_hidden_sheets: Whether to skip hidden Excel sheets.
        file_name: Original file name (required when source is a stream, to detect format).
        color_mode: "ignore" (default) or "value" to fill empty colored cells with hex strings.

    Handles merged cells, error values, hidden rows/cols, and row styles.
    """
    if isinstance(source, str):
        if not os.path.exists(source):
            raise FileReadError(f"File not found: {source}")
        ext = os.path.splitext(source)[1].lower()
        source_label = source
    else:
        if file_name is None:
            raise FileReadError("file_name is required when source is a stream")
        ext = os.path.splitext(file_name)[1].lower()
        source_label = file_name

    if ext in (".xlsx", ".xlsm"):
        return _read_xlsx(source, skip_hidden_sheets, source_label, color_mode)
    elif ext == ".xls":
        return _read_xls(source, skip_hidden_sheets, source_label)
    elif ext in (".csv", ".tsv"):
        return _read_csv(source, ext, source_label)
    else:
        raise UnsupportedFormatError(f"Unsupported format: {ext}")


def _read_xlsx(source, skip_hidden_sheets: bool, source_label: str, color_mode: str = "ignore") -> WorkbookData:
    try:
        wb = openpyxl.load_workbook(source, data_only=True)
    except Exception as e:
        raise FileReadError(f"Cannot read file: {source_label}: {e}") from e

    workbook = WorkbookData(source_file=source_label)

    for ws in wb.worksheets:
        if skip_hidden_sheets and ws.sheet_state != "visible":
            continue

        sheet = SheetData(name=ws.title)

        # Collect merged ranges before reading rows
        merge_ranges = list(ws.merged_cells.ranges)
        merge_map: dict[tuple, Any] = {}
        merge_top_left: dict[tuple, Any] = {}

        for mr in merge_ranges:
            top_left_value = ws.cell(mr.min_row, mr.min_col).value
            for row in range(mr.min_row, mr.max_row + 1):
                for col in range(mr.min_col, mr.max_col + 1):
                    # Store 0-indexed in merge_map
                    merge_map[(row - 1, col - 1)] = top_left_value
                    merge_top_left[(row, col)] = (mr.min_row, mr.min_col)

        sheet.merge_map = merge_map

        # Read all rows
        max_col = ws.max_column or 0
        max_row = ws.max_row or 0
        rows: list[list[Any]] = []

        for row_idx in range(1, max_row + 1):
            row_data: list[Any] = []
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row_idx, col_idx)
                val = cell.value
                # Fill merged cell values
                if val is None and (row_idx - 1, col_idx - 1) in merge_map:
                    val = merge_map[(row_idx - 1, col_idx - 1)]
                # Convert error values to None
                if isinstance(val, str) and val in _ERROR_VALUES:
                    val = None
                # Color-as-data: fill empty cells with hex color if colored
                if val is None and color_mode == "value":
                    rgb = None
                    if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                        r = cell.fill.start_color.rgb
                        if isinstance(r, str) and r != "00000000":
                            rgb = r
                    if rgb:
                        val = f"#{rgb[-6:]}"
                row_data.append(val)
            rows.append(row_data)

        sheet.rows = rows
        sheet.original_col_indices = list(range(max_col))

        # Hidden rows and columns
        hidden_rows = []
        for r in range(1, max_row + 1):
            if ws.row_dimensions[r].hidden:
                hidden_rows.append(r - 1)  # 0-indexed

        hidden_cols = []
        for c in range(1, max_col + 1):
            col_letter = get_column_letter(c)
            if ws.column_dimensions[col_letter].hidden:
                hidden_cols.append(c - 1)  # 0-indexed

        sheet.metadata["hidden_rows"] = hidden_rows
        sheet.metadata["hidden_cols"] = hidden_cols

        # Row styles (bold, fill_color)
        for row_idx in range(1, max_row + 1):
            has_bold = False
            fill_color = None
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row_idx, col_idx)
                if cell.font and cell.font.bold:
                    has_bold = True
                if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                    rgb = cell.fill.start_color.rgb
                    if isinstance(rgb, str) and rgb != "00000000":
                        fill_color = rgb
            sheet.row_styles[row_idx - 1] = {
                "bold": has_bold,
                "fill_color": fill_color,
            }

        workbook.sheets.append(sheet)

    # Uncalculated formula detection (only for file paths, not streams)
    if isinstance(source, str):
        _detect_uncalculated_formulas(source, workbook, skip_hidden_sheets)

    return workbook


def _detect_uncalculated_formulas(
    file_path: str, workbook: WorkbookData, skip_hidden_sheets: bool
) -> None:
    """Check if None values are actually uncalculated formulas."""
    try:
        wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
    except Exception:
        return

    for sheet in workbook.sheets:
        ws = wb_formulas[sheet.name] if sheet.name in wb_formulas.sheetnames else None
        if ws is None:
            continue

        if not sheet.rows:
            continue

        num_cols = len(sheet.rows[0]) if sheet.rows else 0
        if num_cols == 0:
            continue

        warnings = sheet.metadata.get("warnings", [])
        for col_idx in range(num_cols):
            none_count = 0
            formula_count = 0
            for row_idx, row in enumerate(sheet.rows):
                if col_idx < len(row) and row[col_idx] is None:
                    none_count += 1
                    # Check formula in the original workbook
                    cell_val = ws.cell(row_idx + 1, col_idx + 1).value
                    if isinstance(cell_val, str) and cell_val.startswith("="):
                        formula_count += 1

            if none_count > 0 and formula_count / none_count > 0.5:
                warnings.append(
                    f"Column {col_idx} may contain uncalculated formulas "
                    f"({formula_count}/{none_count} None values have formulas)"
                )

        if warnings:
            sheet.metadata["warnings"] = warnings


def _read_xls(source, skip_hidden_sheets: bool, source_label: str) -> WorkbookData:
    """Read .xls files using xlrd (basic support)."""
    try:
        import xlrd
        if isinstance(source, str):
            xls_wb = xlrd.open_workbook(source)
        else:
            xls_wb = xlrd.open_workbook(file_contents=source.read())
    except Exception as e:
        raise FileReadError(f"Cannot read file: {source_label}: {e}") from e

    workbook = WorkbookData(source_file=source_label)

    for sheet_idx in range(xls_wb.nsheets):
        xls_sheet = xls_wb.sheet_by_index(sheet_idx)
        if skip_hidden_sheets and xls_sheet.visibility != 0:
            continue

        sheet = SheetData(name=xls_sheet.name)
        rows: list[list[Any]] = []
        for row_idx in range(xls_sheet.nrows):
            row_data = [xls_sheet.cell_value(row_idx, col) for col in range(xls_sheet.ncols)]
            rows.append(row_data)
        sheet.rows = rows
        sheet.original_col_indices = list(range(xls_sheet.ncols))
        workbook.sheets.append(sheet)

    return workbook


def _read_csv(source, ext: str, source_label: str) -> WorkbookData:
    """Read CSV/TSV files."""
    delimiter = '\t' if ext == '.tsv' else ','

    # Read content from path or stream
    if isinstance(source, str):
        for encoding in ('utf-8', 'utf-8-sig', 'latin-1'):
            try:
                with open(source, 'r', encoding=encoding) as f:
                    content = f.read()
                break
            except UnicodeDecodeError:
                continue
        else:
            raise FileReadError(f"Cannot decode file: {source_label}")
    else:
        raw = source.read()
        if isinstance(raw, bytes):
            for encoding in ('utf-8', 'utf-8-sig', 'latin-1'):
                try:
                    content = raw.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            else:
                raise FileReadError(f"Cannot decode stream: {source_label}")
        else:
            content = raw

    # Try to sniff the delimiter
    try:
        dialect = csv.Sniffer().sniff(content[:8192])
        delimiter = dialect.delimiter
    except csv.Error:
        pass  # use default

    reader = csv.reader(content.splitlines(), delimiter=delimiter)
    rows: list[list[Any]] = []
    for row in reader:
        converted: list[Any] = []
        for cell in row:
            if cell == '':
                converted.append(None)
            else:
                # Try to convert numeric strings so structure detection works
                try:
                    if '.' not in cell and 'e' not in cell.lower():
                        converted.append(int(cell))
                    else:
                        converted.append(float(cell))
                except (ValueError, TypeError):
                    converted.append(cell)
        rows.append(converted)

    sheet_name = os.path.splitext(os.path.basename(source_label))[0]
    sheet = SheetData(name=sheet_name, rows=rows)
    if rows:
        max_col = max(len(r) for r in rows)
        sheet.original_col_indices = list(range(max_col))

    return WorkbookData(source_file=source_label, sheets=[sheet])
