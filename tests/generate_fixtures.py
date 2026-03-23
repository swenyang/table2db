"""Generate all 15 test fixture Excel files with realistic data volumes."""
import os
import random
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

FIXTURES_DIR = os.path.join(os.path.dirname(__file__), "fixtures")


def make_simple():
    """simple.xlsx: Sheet 'Orders', headers + 20 data rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["id", "product", "amount"])
    for i in range(1, 21):
        ws.append([i, f"Product {i}", 10.5 * i])
    wb.save(os.path.join(FIXTURES_DIR, "simple.xlsx"))


def make_merged_cells():
    """merged_cells.xlsx: Merged title, merged group labels, merged subtotal column.

    Headers row 2: id, name, value, group_total
    - Column B: group labels merged across rows (e.g. "Group A" spans 3 rows)
    - Column D: group subtotals merged across rows (e.g. 300 spans 3 rows = sum of column C)
    This tests both label-style and numeric-style merged cells.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    # Row 1: merged title
    ws.merge_cells("A1:D1")
    ws["A1"] = "Sales Report"
    # Row 2: headers
    ws["A2"] = "id"
    ws["B2"] = "name"
    ws["C2"] = "value"
    ws["D2"] = "group_total"

    groups = [
        ("Group A", 3, 5),   # rows 3-5 (1-indexed), 3 rows
        ("Group B", 6, 8),   # rows 6-8, 3 rows
        ("Group C", 9, 11),  # rows 9-11, 3 rows
    ]
    row_id = 1
    for group_name, start_row, end_row in groups:
        group_sum = 0
        for r in range(start_row, end_row + 1):
            val = row_id * 100
            ws.cell(row=r, column=1, value=row_id)
            ws.cell(row=r, column=3, value=val)
            group_sum += val
            row_id += 1
        # Column B: group label merged across rows
        ws.cell(row=start_row, column=2, value=group_name)
        ws.merge_cells(start_row=start_row, start_column=2,
                        end_row=end_row, end_column=2)
        # Column D: group subtotal merged across rows (numeric merge)
        ws.cell(row=start_row, column=4, value=group_sum)
        ws.merge_cells(start_row=start_row, start_column=4,
                        end_row=end_row, end_column=4)

    # Group D: 3 standalone rows (no merge)
    for i in range(3):
        r = 12 + i
        ws.cell(row=r, column=1, value=row_id)
        ws.cell(row=r, column=2, value="Group D")
        ws.cell(row=r, column=3, value=row_id * 100)
        ws.cell(row=r, column=4, value=row_id * 100)  # individual = group total
        row_id += 1

    wb.save(os.path.join(FIXTURES_DIR, "merged_cells.xlsx"))


def make_multi_header():
    """multi_header.xlsx: 2-level merged headers, 20 data rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "MultiHeader"
    # Row 1: top-level headers with merges
    ws["A1"] = "ID"
    ws.merge_cells("A1:A2")
    ws["B1"] = "Personal Info"
    ws.merge_cells("B1:C1")
    ws["D1"] = "Financial"
    ws.merge_cells("D1:E1")
    # Row 2: sub-headers
    ws["B2"] = "Name"
    ws["C2"] = "Age"
    ws["D2"] = "Salary"
    ws["E2"] = "Bonus"
    # Data rows 3-22
    for i in range(1, 21):
        ws.append([i, f"Person_{i}", 21 + i, 39000 + i * 1000, i * 500])
    wb.save(os.path.join(FIXTURES_DIR, "multi_header.xlsx"))


def make_subtotals():
    """subtotals.xlsx: 3 regions × 4 products + subtotals + grand total.

    12 data rows + 3 region subtotals + 1 East "Subtotal" + 1 grand total = 17 body rows.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    bold_font = Font(bold=True)
    ws.append(["region", "product", "quantity", "price"])

    regions = [
        ("North", ["Widget A", "Widget B", "Widget C", "Widget D"]),
        ("South", ["Gadget X", "Gadget Y", "Gadget Z", "Gadget W"]),
        ("East",  ["Part 1", "Part 2", "Part 3", "Part 4"]),
    ]
    current_row = 2  # row 1 is header
    for region, products in regions:
        qty_total = 0
        price_total = 0
        for j, prod in enumerate(products, 1):
            qty = 10 * j
            price = 100 * j
            ws.append([region, prod, qty, price])
            qty_total += qty
            price_total += price
            current_row += 1
        # Chinese subtotal for North/South, English for East
        label = "Subtotal" if region == "East" else "小计"
        ws.append([label, "", qty_total, price_total])
        ws.cell(row=current_row, column=1).font = bold_font
        current_row += 1

    # Grand total
    ws.append(["合计", "", 300, 3000])
    ws.cell(row=current_row, column=1).font = bold_font

    wb.save(os.path.join(FIXTURES_DIR, "subtotals.xlsx"))


def make_mixed_types():
    """mixed_types.xlsx: 25 rows, 20 ints + 5 strings in 'value' column."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Mixed"
    ws.append(["id", "value", "notes"])
    string_positions = {3, 7, 12, 18, 23}  # 1-indexed id positions for strings
    string_values = ["pending", "N/A", "TBD", "unknown", "-"]
    str_idx = 0
    for i in range(1, 26):
        if i in string_positions:
            val = string_values[str_idx]
            str_idx += 1
        else:
            val = i * 100
        ws.append([i, val, f"note {i}"])
    wb.save(os.path.join(FIXTURES_DIR, "mixed_types.xlsx"))


def make_error_values():
    """error_values.xlsx: 20 rows, ~15 numbers + 5 errors in 'calculated'.

    Preserves original pattern: rows[1][1]="#REF!", rows[2][2]="#N/A", rows[3][1]="#DIV/0!".
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Errors"
    ws.append(["id", "calculated", "status"])
    # Row 1 (0-indexed data row 0): keep #REF! in calculated
    ws.append([1, "#REF!", "ok"])
    # Row 2: number in calculated, #N/A in status
    ws.append([2, 100, "#N/A"])
    # Row 3: #DIV/0! in calculated
    ws.append([3, "#DIV/0!", "ok"])
    # Row 4: number, #VALUE! in status
    ws.append([4, 200, "#VALUE!"])
    # Row 5: #NAME? in calculated
    ws.append([5, "#NAME?", "ok"])
    # Rows 6-20: mostly numbers with a couple more errors
    for i in range(6, 21):
        if i == 10:
            ws.append([i, "#REF!", "ok"])
        elif i == 15:
            ws.append([i, "#N/A", "error"])
        else:
            ws.append([i, i * 50, "ok"])
    wb.save(os.path.join(FIXTURES_DIR, "error_values.xlsx"))


def make_offset_table():
    """offset_table.xlsx: title row 1, subtitle row 2, blank row 3, headers row 4, 20 data rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A1"] = "Monthly Sales Report"
    ws["A2"] = "Generated: 2024-01-15"
    # Row 3 empty
    ws["A4"] = "id"
    ws["B4"] = "product"
    ws["C4"] = "revenue"
    for i in range(1, 21):
        ws.cell(row=4 + i, column=1, value=i)
        ws.cell(row=4 + i, column=2, value=f"Item {i}")
        ws.cell(row=4 + i, column=3, value=i * 1000)
    wb.save(os.path.join(FIXTURES_DIR, "offset_table.xlsx"))


def make_empty_gaps():
    """empty_gaps.xlsx: 20 data rows + 2 empty gap rows (after id 7 and id 15)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Gaps"
    ws.append(["id", "name", "score"])
    data_id = 1
    for i in range(1, 23):  # 22 rows total: 20 data + 2 empty
        if i == 8 or i == 16:
            ws.append([None, None, None])
        else:
            ws.append([data_id, f"Person_{data_id}", 60 + data_id])
            data_id += 1
    wb.save(os.path.join(FIXTURES_DIR, "empty_gaps.xlsx"))


def make_multi_sheet_fk():
    """multi_sheet_fk.xlsx: 30 customers + 15 products + 50 orders."""
    wb = Workbook()
    # Customers
    ws1 = wb.active
    ws1.title = "Customers"
    ws1.append(["customer_id", "name", "email"])
    for i in range(1, 31):
        ws1.append([i, f"Customer {i}", f"customer{i}@example.com"])
    # Products
    ws2 = wb.create_sheet("Products")
    ws2.append(["product_id", "product_name", "price"])
    for i in range(1, 16):
        ws2.append([i, f"Product {i}", round(10.0 + i * 5.5, 2)])
    # Orders
    ws3 = wb.create_sheet("Orders")
    ws3.append(["order_id", "customer_id", "product_id", "quantity"])
    random.seed(42)
    for i in range(1, 51):
        ws3.append([
            i,
            random.randint(1, 30),
            random.randint(1, 15),
            random.randint(1, 50),
        ])
    wb.save(os.path.join(FIXTURES_DIR, "multi_sheet_fk.xlsx"))


def make_number_as_text():
    """number_as_text.xlsx: 20 rows with numbers stored as strings."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Numbers"
    ws.append(["id", "code", "amount"])
    for i in range(1, 21):
        ws.append([i, str(2000 + i), str(round(i * 12.75, 2))])
    wb.save(os.path.join(FIXTURES_DIR, "number_as_text.xlsx"))


def make_dates_mixed():
    """dates_mixed.xlsx: 20 rows alternating datetime objects and date strings."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dates"
    ws.append(["id", "event_date", "label"])
    base = datetime.datetime(2024, 1, 1)
    for i in range(1, 21):
        if i % 2 == 1:
            # Odd rows: datetime object
            dt = base + datetime.timedelta(days=i * 15)
            ws.append([i, dt, f"Event_{i}"])
        else:
            # Even rows: date string, alternate formats
            dt = base + datetime.timedelta(days=i * 15)
            if i % 4 == 0:
                date_str = dt.strftime("%Y/%m/%d")
            else:
                date_str = dt.strftime("%Y-%m-%d")
            ws.append([i, date_str, f"Event_{i}"])
    wb.save(os.path.join(FIXTURES_DIR, "dates_mixed.xlsx"))


def make_hidden_rows_cols():
    """hidden_rows_cols.xlsx: 20 data rows with hidden rows and a hidden column.

    Uses numeric-heavy data to avoid false multi-header detection.
    Hides column D ('notes'), and rows 5, 10, 15 (1-indexed = data rows 4, 9, 14).
    All 20 data rows including hidden ones should appear in the DB.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    ws.append(["id", "name", "salary", "notes", "department"])
    for i in range(1, 21):
        ws.append([
            i,
            f"Employee {i}",
            30000 + i * 1000,
            f"internal note {i}",
            ["Engineering", "Sales", "Marketing", "HR"][i % 4],
        ])
    # Hide column D (notes - internal data)
    ws.column_dimensions["D"].hidden = True
    # Hide rows 5, 10, 15 (1-indexed; data rows for id=4, id=9, id=14)
    ws.row_dimensions[5].hidden = True
    ws.row_dimensions[10].hidden = True
    ws.row_dimensions[15].hidden = True
    wb.save(os.path.join(FIXTURES_DIR, "hidden_rows_cols.xlsx"))


def make_empty_after_clean():
    """empty_after_clean.xlsx: Few data rows + many subtotal/total rows.

    Realistic scenario: a small summary table where most rows are aggregations.
    3 data rows + 4 total rows. After cleaning → 3 data rows survive.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    bold = Font(bold=True)
    ws.append(["category", "region", "amount"])
    ws.append(["Electronics", "North", 1200])
    ws.append(["Electronics", "South", 800])
    ws.append(["小计", "", 2000])
    ws.cell(row=4, column=1).font = bold
    ws.append(["Furniture", "North", 3000])
    ws.append(["小计", "", 3000])
    ws.cell(row=6, column=1).font = bold
    ws.append(["合计", "", 5000])
    ws.cell(row=7, column=1).font = bold
    wb.save(os.path.join(FIXTURES_DIR, "empty_after_clean.xlsx"))


def make_duplicate_sheet_names():
    """duplicate_sheet_names.xlsx: Two sheets with similar names, 15 rows each."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sales Data"
    ws1.append(["id", "region", "total"])
    regions = ["North", "South", "East", "West", "Central"]
    for i in range(1, 16):
        ws1.append([i, regions[i % len(regions)], i * 100])

    ws2 = wb.create_sheet("Sales-Data")
    ws2.append(["id", "product", "qty"])
    products = ["Widget", "Gadget", "Doohickey", "Thingamajig", "Gizmo"]
    for i in range(1, 16):
        ws2.append([i, products[i % len(products)], i * 10])
    wb.save(os.path.join(FIXTURES_DIR, "duplicate_sheet_names.xlsx"))


def make_real_world_dirty():
    """real_world_dirty.xlsx: 4 regions × 5 products + subtotals + grand total.

    20 data rows + 4 region subtotals + 1 grand total = 25 body rows.
    One "pending" value in Amount column (19/20 = 95% numeric).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Q1 Report"
    bold_font = Font(bold=True)

    # Row 1: merged title
    ws.merge_cells("A1:E1")
    ws["A1"] = "Q1 2024 Sales Report"
    ws["A1"].font = Font(bold=True, size=14)
    # Row 2: subtitle
    ws["A2"] = "Generated on 2024-04-01"
    # Row 3: empty
    # Rows 4-5: multi-level headers
    ws["A4"] = "Region"
    ws.merge_cells("A4:A5")
    ws["B4"] = "Product Details"
    ws.merge_cells("B4:C4")
    ws["D4"] = "Financials"
    ws.merge_cells("D4:E4")
    ws["B5"] = "Name"
    ws["C5"] = "Category"
    ws["D5"] = "Amount"
    ws["E5"] = "Tax"

    regions = [
        ("North", "Electronics", ["Widget A", "Widget B", "Widget C", "Widget D", "Widget E"]),
        ("South", "Hardware",    ["Gadget X", "Gadget Y", "Gadget Z", "Gadget W", "Gadget V"]),
        ("East",  "Software",    ["App 1", "App 2", "App 3", "App 4", "App 5"]),
        ("West",  "Services",    ["Service A", "Service B", "Service C", "Service D", "Service E"]),
    ]
    current_row = 6
    product_idx = 0
    for region, category, products in regions:
        amount_total = 0
        tax_total = 0
        for j, prod in enumerate(products):
            product_idx += 1
            amount = product_idx * 500
            tax = product_idx * 50
            # One "pending" value for South/Gadget Y (product_idx == 7)
            if product_idx == 7:
                amount = "pending"
            for c_idx, val in enumerate(
                [region, prod, category, amount, tax], 1
            ):
                ws.cell(row=current_row, column=c_idx, value=val)
            if isinstance(amount, (int, float)):
                amount_total += amount
                tax_total += tax
            else:
                tax_total += tax
            current_row += 1
        # Region subtotal (bold)
        subtotal_label = f"{region} Subtotal"
        for c_idx, val in enumerate(
            [subtotal_label, "", "", amount_total, tax_total], 1
        ):
            ws.cell(row=current_row, column=c_idx, value=val)
        ws.cell(row=current_row, column=1).font = bold_font
        current_row += 1

    # Grand total
    for c_idx, val in enumerate(["Grand Total", "", "", 99999, 9999], 1):
        ws.cell(row=current_row, column=c_idx, value=val)
    ws.cell(row=current_row, column=1).font = bold_font

    wb.save(os.path.join(FIXTURES_DIR, "real_world_dirty.xlsx"))


def make_multi_table_sheet():
    """multi_table_sheet.xlsx: One sheet with two separate tables separated by empty rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined"

    # Table 1: Employee data (rows 1-11, 1 header + 10 data)
    ws.append(["emp_id", "name", "department", "salary"])
    for i in range(1, 11):
        ws.append([i, f"Employee {i}", ["Engineering", "Sales", "HR"][i % 3], 40000 + i * 2000])

    # Gap: 3 empty rows
    ws.append([None, None, None, None])
    ws.append([None, None, None, None])
    ws.append([None, None, None, None])

    # Table 2: Project data (1 header + 10 data)
    ws.append(["project_id", "project_name", "budget"])
    for i in range(1, 11):
        ws.append([100 + i, f"Project {chr(64+i)}", i * 50000])

    wb.save(os.path.join(FIXTURES_DIR, "multi_table_sheet.xlsx"))


def make_csv_fixture():
    """students_csv.csv: basic CSV file with student scores."""
    import csv as csv_mod
    path = os.path.join(FIXTURES_DIR, "students_csv.csv")
    with open(path, 'w', newline='', encoding='utf-8') as f:
        writer = csv_mod.writer(f)
        writer.writerow(["id", "name", "score"])
        for i in range(1, 21):
            writer.writerow([i, f"Student {i}", 60 + i])


def make_tsv_fixture():
    """products_tsv.tsv: basic TSV file with product prices."""
    path = os.path.join(FIXTURES_DIR, "products_tsv.tsv")
    with open(path, 'w', newline='', encoding='utf-8') as f:
        f.write("id\tproduct\tprice\n")
        for i in range(1, 16):
            f.write(f"{i}\tItem {i}\t{i * 9.99:.2f}\n")


def main():
    os.makedirs(FIXTURES_DIR, exist_ok=True)
    generators = [
        make_simple,
        make_merged_cells,
        make_multi_header,
        make_subtotals,
        make_mixed_types,
        make_error_values,
        make_offset_table,
        make_empty_gaps,
        make_multi_sheet_fk,
        make_number_as_text,
        make_dates_mixed,
        make_hidden_rows_cols,
        make_empty_after_clean,
        make_duplicate_sheet_names,
        make_real_world_dirty,
        make_multi_table_sheet,
        make_csv_fixture,
        make_tsv_fixture,
    ]
    for gen in generators:
        gen()
    print(f"Generated {len(generators)} fixture files in {FIXTURES_DIR}")


if __name__ == "__main__":
    main()
